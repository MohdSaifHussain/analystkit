"""analystkit.workpaper — analysis packaged as re-performable evidence.

Atomic writes: tempfile.mkstemp + os.replace, so a crash mid-write can
never leave a partial workpaper on disk (os.replace is atomic per the
Python official documentation of os.replace).
"""
from __future__ import annotations

import os
import tempfile
from collections.abc import Callable
from datetime import datetime
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analystkit.core import IST, PAL, Dimension
from analystkit.dedupe import find_duplicates
from analystkit.profiling import dimension_scores, profile_columns
from analystkit.rules import RuleResult, load_rules, run_rules
from analystkit.teach import LESSONS

# ─────────────────────────────────────────────────────────────────────────────

def _fill(colour: str) -> PatternFill:
    return PatternFill("solid", start_color=colour)


def _font(bold: bool = False, colour: str = PAL["ink"],
          size: int = 10, italic: bool = False) -> Font:
    return Font(name="Arial", bold=bold, color=colour, size=size, italic=italic)


def _head(ws: Worksheet, headers: list[str], row: int) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color=PAL["white"], size=10)
        cell.fill = _fill(PAL["navy"])
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def _fit(ws: Worksheet, mn: int = 10, mx: int = 70) -> None:
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None),
                    default=mn)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(width + 2, mn), mx)


def _sheet_rows(ws: Worksheet, headers: list[str],
                rows: list[tuple[Any, ...]], start: int = 3) -> None:
    _head(ws, headers, start)
    for r, row in enumerate(rows, start + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = _font()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 0:
                cell.fill = _fill(PAL["alt"])
    ws.freeze_panes = ws.cell(row=start + 1, column=1).coordinate
    _fit(ws)


def cmd_workpaper(
    source: str, rules_path: Path | None, key: str | None,
    out: Path | None, table: str | None,
    opener: Callable[[str, str | None], duckdb.DuckDBPyConnection],
) -> None:
    """Runs profile + dedupe (+ validate when rules given) and packages
    everything into a reviewer-grade Excel workpaper. Atomic write."""
    con = opener(source, table)
    profiles = profile_columns(con)
    scores = dimension_scores(con, profiles)
    dup_rows, dup_groups = find_duplicates(con, key)
    rule_results: list[RuleResult] = []
    if rules_path:
        rule_results = run_rules(con, load_rules(rules_path))

    run_at = datetime.now(tz=IST)
    out_path = out or Path(f"workpaper_{run_at.strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb = Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)

    # Tab 1: Methodology (the re-performance record)
    ws = wb.create_sheet("Methodology")
    ws["A1"] = "DATA QUALITY WORKPAPER"
    ws["A1"].font = _font(bold=True, size=15, colour=PAL["navy"])
    method = [
        ("Prepared by", "AnalystKit v1.1 (operator: Mohd Saif Hussain)"),
        ("Run timestamp", run_at.strftime("%d %B %Y, %H:%M IST")),
        ("Source", str(source)),
        ("Rows in scope", f"{profiles[0].total if profiles else 0:,}"),
        ("Framework", "DAMA-DMBOK six data quality dimensions"),
        ("Procedures", "1) Column profile  2) Dimension scoring  "
                       "3) Duplicate detection"
                       + ("  4) Rule validation" if rule_results else "")),
        ("Duplicate scope", f"key '{key}'" if key else "entire row"),
        ("Rules file", str(rules_path) if rules_path else "none supplied"),
        ("Re-performance", "Every figure derives from SQL over the stated "
                           "source; re-running the same commands on the same "
                           "file reproduces this workpaper exactly."),
        ("Limitation", "Accuracy is not scored: it requires reconciliation "
                       "against an authoritative source (see reconcile "
                       "command). Scores reflect the dataset as provided."),
    ]
    for i, (label, value) in enumerate(method, start=3):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = _font()
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90

    # Tab 2: DQ Scorecard
    ws2 = wb.create_sheet("DQ Scorecard")
    ws2["A1"] = "DAMA SIX-DIMENSION SCORECARD"
    ws2["A1"].font = _font(bold=True, size=13, colour=PAL["navy"])
    score_rows: list[tuple[Any, ...]] = []
    for dim, s in scores.items():
        if s is None:
            note = ("Requires reconcile against authoritative source"
                    if dim is Dimension.ACCURACY else "No time column detected")
            score_rows.append((str(dim), "not scored", note))
        else:
            rating = "STRONG" if s >= 0.95 else ("REVIEW" if s >= 0.85 else "WEAK")
            score_rows.append((str(dim), f"{s * 100:.1f}%", rating))
    _sheet_rows(ws2, ["Dimension", "Score", "Assessment"], score_rows)
    for r in range(4, 4 + len(score_rows)):
        v = str(ws2.cell(row=r, column=3).value)
        cell = ws2.cell(row=r, column=3)
        if v == "STRONG":
            cell.fill = _fill(PAL["green"])
            cell.font = _font(bold=True, colour=PAL["white"])
        elif v == "REVIEW":
            cell.fill = _fill(PAL["amber"])
            cell.font = _font(bold=True)
        elif v == "WEAK":
            cell.fill = _fill(PAL["red"])
            cell.font = _font(bold=True, colour=PAL["white"])

    # Tab 3: Column Profile
    ws3 = wb.create_sheet("Column Profile")
    ws3["A1"] = "PER-COLUMN PROFILE"
    ws3["A1"].font = _font(bold=True, size=13, colour=PAL["navy"])
    _sheet_rows(
        ws3,
        ["Column", "Type", "Nulls", "Null %", "Distinct", "Case variants", "Valid %"],
        [(p.name, p.dtype, p.nulls,
          f"{(p.nulls / p.total * 100 if p.total else 0):.1f}%",
          p.distinct, p.case_variants, f"{p.valid_ratio * 100:.1f}%")
         for p in profiles],
    )

    # Tab 4: Findings
    ws4 = wb.create_sheet("Findings")
    ws4["A1"] = "FINDINGS — EXCEPTIONS AND DUPLICATES"
    ws4["A1"].font = _font(bold=True, size=13, colour=PAL["navy"])
    finding_rows: list[tuple[Any, ...]] = [
        ("DUP-01", "Duplicates",
         f"{len(dup_groups):,} duplicate groups; {dup_rows:,} excess rows "
         f"({'key: ' + key if key else 'full-row'})",
         "Investigate source of duplication before any aggregation.")
    ]
    for rr in rule_results:
        finding_rows.append((
            rr.rule_id, f"{rr.column} · {rr.rule}",
            f"{rr.failures:,} exceptions — {rr.detail}"
            + (f" | sample: {'; '.join(rr.sample)}" if rr.sample else ""),
            "Remediate at source; exceptions reported, never dropped.",
        ))
    _sheet_rows(ws4, ["ID", "Area", "Finding", "Recommended action"], finding_rows)

    # Tab 5: Learn (the self-teaching layer)
    ws5 = wb.create_sheet("Learn")
    ws5["A1"] = "LEARN — WHAT EACH DIMENSION MEANS AND THE SQL BEHIND IT"
    ws5["A1"].font = _font(bold=True, size=13, colour=PAL["navy"])
    _sheet_rows(
        ws5, ["Topic", "Lesson"],
        [(topic.title(), LESSONS[topic]) for topic in
         ("completeness", "uniqueness", "validity", "consistency",
          "timeliness", "accuracy", "reconcile", "workpaper")],
    )
    ws5.column_dimensions["A"].width = 16
    ws5.column_dimensions["B"].width = 100

    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=out_path.parent or Path("."))
    os.close(fd)
    tmp_path = Path(tmp)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, out_path)
    except BaseException:
        tmp_path.unlink(missing_ok=True)
        raise
    print(f"  ✓  Workpaper ready: {out_path}")
    print("     Tabs: Methodology · DQ Scorecard · Column Profile · "
          "Findings · Learn")

